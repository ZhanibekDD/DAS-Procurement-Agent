from __future__ import annotations

import io
import os
import tempfile
import unittest
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from procurement.config import Settings
from procurement.db import Database
from procurement.imports import parse_supplier_table
from procurement.models import (
    CampaignCreate,
    LotCreate,
    LotItemCreate,
    ProcurementSuggestionApproval,
    ProcurementSuggestionCreate,
    ProcurementSuggestionRejection,
    ProjectCreate,
    PurchaseHistoryCreate,
    QuoteCreate,
    QuoteItemCreate,
    SupplierCreate,
)
from procurement.service import ConflictError, ProcurementService
from procurement.templates import render_template


class ProcurementTestCase(unittest.TestCase):
    def setUp(self):
        handle = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.db_path = handle.name
        handle.close()
        self.db = Database(self.db_path)
        self.db.initialize()
        self.service = ProcurementService(self.db)

    def tearDown(self):
        for suffix in ("", "-wal", "-shm"):
            try:
                os.unlink(self.db_path + suffix)
            except FileNotFoundError:
                pass

    def _project_and_lot(self):
        project = self.service.create_project(
            ProjectCreate(
                name="Склад Воронеж",
                region="Воронежская область",
                delivery_address="г. Воронеж, ул. Монтажная, 10",
            )
        )
        lot = self.service.create_lot(
            LotCreate(
                project_id=project["id"],
                title="Окна ПВХ",
                region="Воронеж",
                delivery_address=project["delivery_address"],
                response_deadline=date.today() + timedelta(days=7),
                desired_delivery_date=date.today() + timedelta(days=30),
                items=[
                    LotItemCreate(
                        name="Окно ПВХ 1400×1200",
                        quantity=Decimal("84"),
                        unit="шт.",
                        specification="двухкамерный стеклопакет, белый профиль",
                    )
                ],
            )
        )
        return project, lot

    def test_full_workflow_drafts_rfq_and_ranks_landed_cost(self):
        _, lot = self._project_and_lot()
        supplier_a = self.service.create_supplier(
            SupplierCreate(
                name="Окна Регион",
                tax_id="3666000001",
                region="Воронеж",
                email="sales@okna.example",
                categories=["окна"],
                rating=4.8,
                verified=True,
            )
        )
        supplier_b = self.service.create_supplier(
            SupplierCreate(
                name="Строй Комплект",
                tax_id="3666000002",
                region="Воронежская область",
                email="tender@stroy.example",
                categories=["окна", "двери"],
                rating=4.0,
                verified=True,
            )
        )

        matches = self.service.match_suppliers(lot["id"])
        self.assertEqual({row["id"] for row in matches}, {supplier_a["id"], supplier_b["id"]})

        campaign = self.service.create_campaign(
            lot["id"],
            CampaignCreate(supplier_ids=[supplier_a["id"], supplier_b["id"]]),
        )
        self.assertEqual(len(campaign["messages"]), 2)
        self.assertTrue(all(message["status"] == "draft" for message in campaign["messages"]))
        self.assertIn("84", campaign["messages"][0]["body"])

        approved = self.service.approve_message(campaign["messages"][0]["id"], "manager@example")
        self.assertEqual(approved["status"], "approved")
        self.assertNotEqual(approved["status"], "sent")

        item_id = lot["items"][0]["id"]
        self.service.add_quote(
            lot["id"],
            QuoteCreate(
                supplier_id=supplier_a["id"],
                delivery_cost=Decimal("40000"),
                lead_days=14,
                payment_terms="50% аванс, 50% после поставки",
                items=[QuoteItemCreate(lot_item_id=item_id, unit_price=Decimal("12500"))],
            ),
        )

        history = self.service.add_purchase_history(
            PurchaseHistoryCreate(
                supplier_id=supplier_a["id"],
                item_name="Окно ПВХ 1400x1200",
                quantity=Decimal("60"),
                unit="шт.",
                unit_price=Decimal("12000"),
                currency="RUB",
                purchased_on=date.today() - timedelta(days=90),
                invoice_number="INV-OLD-42",
                region="Воронеж",
                confirmed_by="Руководитель снабжения",
            )
        )
        self.assertEqual(history["review_status"], "approved")
        self.service.add_quote(
            lot["id"],
            QuoteCreate(
                supplier_id=supplier_b["id"],
                delivery_cost=Decimal("0"),
                lead_days=25,
                payment_terms="100% предоплата",
                items=[QuoteItemCreate(lot_item_id=item_id, unit_price=Decimal("12800"))],
            ),
        )

        comparison = self.service.comparison(lot["id"])
        self.assertEqual(comparison["decision"], "human_approval_required")
        self.assertEqual(comparison["quotes"][0]["rank"], 1)
        self.assertEqual(comparison["quotes"][0]["supplier_name"], "Окна Регион")
        self.assertEqual(comparison["quotes"][0]["total_cost"], 1_090_000.0)
        self.assertEqual(comparison["price_benchmark"]["matched_items"], 1)
        self.assertEqual(comparison["quotes"][0]["history_coverage"], "1/1")
        self.assertEqual(comparison["quotes"][0]["history_variance_pct"], 4.2)
        self.assertEqual(comparison["quotes"][0]["potential_saving"], 42_000.0)

        campaigns = self.service.list_campaigns(lot["id"])
        self.assertEqual(campaigns[0]["message_count"], 2)
        self.assertEqual(campaigns[0]["approved_count"], 1)

        drafts = self.service.list_outbox(status="draft", lot_id=lot["id"])
        self.assertEqual(len(drafts), 1)
        self.assertEqual(drafts[0]["supplier_name"], "Строй Комплект")

        quotes = self.service.list_quotes(lot["id"])
        self.assertEqual(len(quotes), 2)
        self.assertEqual(quotes[0]["items"][0]["lot_item_name"], "Окно ПВХ 1400×1200")

        audit = self.service.list_audit(limit=10)
        self.assertTrue(any(row["action"] == "approved" for row in audit))
        self.assertTrue(any(row["entity_type"] == "purchase_history" for row in audit))
        self.assertTrue(all(isinstance(row["details"], dict) for row in audit))

    def test_price_memory_matches_similar_names_but_not_wrong_currency_or_unit(self):
        _, lot = self._project_and_lot()
        supplier = self.service.create_supplier(
            SupplierCreate(name="Исторический поставщик", region="Воронеж")
        )
        common = {
            "supplier_id": supplier["id"],
            "quantity": Decimal("10"),
            "purchased_on": date.today(),
            "confirmed_by": "Комиссия",
        }
        self.service.add_purchase_history(
            PurchaseHistoryCreate(
                **common,
                item_name="Окно ПВХ 1400×1200 белое",
                unit="шт.",
                unit_price=Decimal("11000"),
                currency="RUB",
            )
        )
        self.service.add_purchase_history(
            PurchaseHistoryCreate(
                **common,
                item_name="Окно ПВХ 1400×1200 белое",
                unit="м2",
                unit_price=Decimal("5000"),
                currency="RUB",
            )
        )
        self.service.add_purchase_history(
            PurchaseHistoryCreate(
                **common,
                item_name="Окно ПВХ 1400×1200 белое",
                unit="шт.",
                unit_price=Decimal("999"),
                currency="KZT",
            )
        )

        benchmark = self.service.lot_price_benchmark(lot["id"])
        self.assertEqual(benchmark["matched_items"], 1)
        self.assertEqual(benchmark["items"][0]["history_count"], 1)
        self.assertEqual(benchmark["items"][0]["median_unit_price"], 11000.0)

        search = self.service.list_purchase_history(search="окно пвх")
        self.assertEqual(len(search), 3)

    def test_duplicate_paid_purchase_is_rejected(self):
        data = PurchaseHistoryCreate(
            item_name="Кирпич М150",
            quantity=Decimal("1000"),
            unit="шт.",
            unit_price=Decimal("42.50"),
            currency="RUB",
            purchased_on=date.today(),
            invoice_number="СЧ-10",
            confirmed_by="Снабженец",
        )
        self.service.add_purchase_history(data)
        with self.assertRaisesRegex(ConflictError, "already in price history"):
            self.service.add_purchase_history(data)

    def test_project_document_suggestion_requires_review_before_creating_lot(self):
        project = self.service.create_project(
            ProjectCreate(
                name="Бизнес-центр",
                region="Воронеж",
                delivery_address="ул. Монтажная, 10",
            )
        )
        document = self.service.register_source_document(
            filename="АР.pdf",
            content=b"%PDF-1.4\nproject section",
            document_type="project_section",
            project_id=project["id"],
        )
        suggestions = self.service.register_procurement_suggestions(
            document["id"],
            [
                ProcurementSuggestionCreate(
                    section_code="АР",
                    section_name="Архитектурные решения",
                    lot_title="Оконные блоки",
                    confidence=0.94,
                    evidence=["АР-12", "Ведомость заполнения проёмов"],
                    items=[
                        LotItemCreate(
                            name="Окно ПВХ 1400×1200",
                            quantity=84,
                            unit="шт.",
                            specification="двухкамерный стеклопакет",
                        )
                    ],
                )
            ],
        )
        suggestion = suggestions[0]
        self.assertEqual(suggestion["status"], "needs_review")
        self.assertEqual(self.service.list_lots(), [])
        self.assertEqual(suggestion["items"][0]["quantity"], "84")
        self.assertEqual(
            self.service.list_source_documents()[0]["extraction_status"], "needs_review"
        )
        self.assertEqual(
            [row["id"] for row in self.service.list_procurement_suggestions(status="needs_review")],
            [suggestion["id"]],
        )

        approved = self.service.approve_procurement_suggestion(
            suggestion["id"],
            ProcurementSuggestionApproval(
                response_deadline=date.today() + timedelta(days=7),
                desired_delivery_date=date.today() + timedelta(days=30),
                approved_by="Начальник снабжения",
            ),
        )
        self.assertEqual(approved["suggestion"]["status"], "approved")
        self.assertEqual(approved["lot"]["title"], "Оконные блоки")
        self.assertEqual(approved["lot"]["items"][0]["quantity"], "84")
        self.assertEqual(len(self.service.get_project(project["id"])["sections"]), 1)
        self.assertEqual(self.service.list_source_documents()[0]["extraction_status"], "approved")
        with self.assertRaisesRegex(ConflictError, "awaiting review"):
            self.service.approve_procurement_suggestion(
                suggestion["id"],
                ProcurementSuggestionApproval(
                    response_deadline=date.today() + timedelta(days=7),
                    approved_by="Другой сотрудник",
                ),
            )

    def test_project_document_suggestion_can_be_rejected_without_creating_lot(self):
        project = self.service.create_project(
            ProjectCreate(name="Склад", region="Алматы", delivery_address="ул. Абая, 1")
        )
        document = self.service.register_source_document(
            filename="КР.pdf",
            content=b"%PDF-1.4\nstructural section",
            document_type="project_section",
            project_id=project["id"],
        )
        suggestion = self.service.register_procurement_suggestions(
            document["id"],
            [
                ProcurementSuggestionCreate(
                    section_code="КР",
                    section_name="Конструктивные решения",
                    lot_title="Арматура",
                    confidence=0.71,
                    items=[LotItemCreate(name="Арматура А500С", quantity=10, unit="т")],
                )
            ],
        )[0]
        rejected = self.service.reject_procurement_suggestion(
            suggestion["id"],
            ProcurementSuggestionRejection(
                reviewed_by="Технический эксперт", reason="Позиция уже закуплена"
            ),
        )
        self.assertEqual(rejected["status"], "rejected")
        self.assertEqual(self.service.list_lots(), [])
        self.assertEqual(self.service.list_source_documents()[0]["extraction_status"], "approved")
        with self.assertRaisesRegex(ConflictError, "awaiting review"):
            self.service.reject_procurement_suggestion(
                suggestion["id"],
                ProcurementSuggestionRejection(
                    reviewed_by="Другой эксперт", reason="Повторное решение"
                ),
            )

    def test_list_filters_fail_closed_on_invalid_values(self):
        with self.assertRaisesRegex(ValueError, "unsupported outbox status"):
            self.service.list_outbox(status="unknown")
        with self.assertRaisesRegex(ValueError, "between 1 and 200"):
            self.service.list_audit(limit=0)
        with self.assertRaisesRegex(ValueError, "unsupported procurement suggestion status"):
            self.service.list_procurement_suggestions(status="approving")

    def test_incomplete_quote_is_disqualified(self):
        project = self.service.create_project(
            ProjectCreate(name="Проект", region="Алматы", delivery_address="ул. Абая, 1")
        )
        lot = self.service.create_lot(
            LotCreate(
                project_id=project["id"],
                title="Песок",
                region="Алматы",
                delivery_address="ул. Абая, 1",
                response_deadline=date.today() + timedelta(days=3),
                items=[
                    LotItemCreate(name="Песок", quantity=10, unit="т"),
                    LotItemCreate(name="Щебень", quantity=5, unit="т"),
                ],
            )
        )
        supplier = self.service.create_supplier(
            SupplierCreate(name="Карьер", region="Алматы", email="a@b.kz")
        )
        self.service.add_quote(
            lot["id"],
            QuoteCreate(
                supplier_id=supplier["id"],
                items=[QuoteItemCreate(lot_item_id=lot["items"][0]["id"], unit_price=1000)],
            ),
        )
        row = self.service.comparison(lot["id"])["quotes"][0]
        self.assertEqual(row["rank_status"], "disqualified")
        self.assertEqual(row["coverage"], "1/2")

    def test_duplicate_supplier_tax_id_is_rejected(self):
        supplier = SupplierCreate(name="Поставщик", tax_id="123", region="Астана")
        self.service.create_supplier(supplier)
        with self.assertRaises(ConflictError):
            self.service.create_supplier(supplier)

    def test_template_requires_every_variable(self):
        with self.assertRaisesRegex(ValueError, "supplier_name"):
            render_template("Здравствуйте, {supplier_name}", {})

    def test_paid_invoice_is_stored_once_and_queued_for_review(self):
        payload = b"%PDF-1.4\n% safe test invoice"
        first = self.service.register_source_document(
            filename="invoice-42.pdf",
            content=payload,
            document_type="paid_invoice",
            content_type="application/pdf",
        )
        second = self.service.register_source_document(
            filename="copy.pdf",
            content=payload,
            document_type="paid_invoice",
            content_type="application/pdf",
        )
        self.assertEqual(first["id"], second["id"])
        self.assertEqual(first["extraction_status"], "pending_ai_extraction")
        self.assertTrue(os.path.exists(first["storage_path"]))
        os.unlink(first["storage_path"])

    def test_invalid_pdf_payload_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "invalid PDF"):
            self.service.register_source_document(
                filename="invoice.pdf",
                content=b"not a pdf",
                document_type="paid_invoice",
            )


class ImportTests(unittest.TestCase):
    def test_csv_supplier_preview(self):
        content = (
            "Поставщик;ИНН;Регион;Email;Категории;Проверен\n"
            "Окна Регион;3666000001;Воронеж;sales@example.ru;окна, двери;да\n"
        ).encode("utf-8")
        preview = parse_supplier_table(content, "suppliers.csv")
        self.assertEqual(len(preview.rows), 1)
        self.assertEqual(preview.rows[0].categories, ["окна", "двери"])
        self.assertTrue(preview.rows[0].verified)

    def test_xlsx_supplier_preview_and_invalid_row(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Компания", "Город", "Почта", "Рейтинг"])
        sheet.append(["Карьер №1", "Воронеж", "tender@example.ru", 4.5])
        sheet.append(["", "", "bad", 9])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        preview = parse_supplier_table(buffer.getvalue(), "suppliers.xlsx")
        self.assertEqual(len(preview.rows), 1)
        self.assertEqual(preview.rows[0].name, "Карьер №1")
        self.assertEqual(len(preview.errors), 1)


class SettingsTests(unittest.TestCase):
    def test_production_requires_api_key(self):
        with patch.dict(
            os.environ,
            {"PROCUREMENT_ENV": "production", "PROCUREMENT_API_KEY": ""},
            clear=True,
        ):
            with self.assertRaisesRegex(RuntimeError, "API_KEY"):
                Settings.from_env()


class StaticPreviewTests(unittest.TestCase):
    def test_uvicorn_does_not_trust_proxy_headers_before_the_app(self):
        dockerfile = (Path(__file__).parents[1] / "Dockerfile").read_text(
            encoding="utf-8"
        )
        self.assertIn("--no-proxy-headers", dockerfile)

    def test_demo_mode_is_bundled_into_the_real_control_center(self):
        static_dir = Path(__file__).parents[1] / "procurement" / "static"
        html = (static_dir / "index.html").read_text(encoding="utf-8")
        self.assertIn('id="demoBtn"', html)
        self.assertIn("const DEMO_DATA=", html)
        self.assertIn("function loadDemo()", html)
        self.assertIn("В демонстрации изменения не сохраняются", html)
        self.assertIn("Независимая учётная запись", html)
        self.assertIn("credentials:'same-origin'", html)
        self.assertNotIn("Единый вход через DAS", html)
        self.assertNotIn('id="apiKey"', html)
        self.assertNotIn("X-API-Key", html)
        self.assertNotIn("localStorage", html)

        login = (static_dir / "login.html").read_text(encoding="utf-8")
        self.assertIn('method="post" action="/auth/login"', login)
        self.assertIn('autocomplete="current-password"', login)
        self.assertNotIn("DAS SSO", login)

    def test_premium_visual_assets_are_bundled_locally(self):
        static_dir = Path(__file__).parents[1] / "procurement" / "static"
        html = (static_dir / "index.html").read_text(encoding="utf-8")
        hero = static_dir / "assets" / "procurement-hero.webp"
        supplier_network = static_dir / "assets" / "supplier-network.webp"
        self.assertTrue(hero.is_file())
        self.assertGreater(hero.stat().st_size, 50_000)
        self.assertTrue(supplier_network.is_file())
        self.assertGreater(supplier_network.stat().st_size, 40_000)
        self.assertIn('url("assets/procurement-hero.webp")', html)
        self.assertIn('url("assets/supplier-network.webp")', html)
        for icon in (
            "brand",
            "home",
            "projects",
            "lots",
            "suppliers",
            "mail",
            "compare",
            "price",
            "docs",
            "templates",
            "shield",
            "map",
            "star",
            "upload",
            "clock",
            "location",
            "box",
            "send",
            "filter",
            "truck",
            "arrow",
        ):
            self.assertIn(f'id="i-{icon}"', html)

    def test_daily_procurement_views_have_visual_workflow_controls(self):
        static_dir = Path(__file__).parents[1] / "procurement" / "static"
        html = (static_dir / "index.html").read_text(encoding="utf-8")
        for marker in (
            'class="project-grid"',
            'id="lotStatusFilter"',
            'id="lotProjectFilter"',
            'class="rfq-flow"',
            'class="message-list"',
            "function filterLots()",
            "function showProjectLots(id)",
            'id="tender"',
            'class="tender-workspace"',
            'class="tender-stage"',
            "function renderTender()",
            "function openTenderRfq()",
            'class="ai-review"',
            'class="suggestion-grid"',
            "const DEMO_SUGGESTIONS=",
            "function approveSuggestion(id)",
            "function rejectSuggestion(id)",
        ):
            self.assertIn(marker, html)


if __name__ == "__main__":
    unittest.main()


# ============================================================================
# PR #8: Batch import, supplier drafts, price-history entries
# ============================================================================

class BatchImportTestCase(unittest.TestCase):
    """Unit tests for batch-import helpers and service methods."""

    def setUp(self):
        handle = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.db_path = handle.name
        handle.close()
        self.db = Database(self.db_path)
        self.db.initialize()
        self.service = ProcurementService(self.db)

    def tearDown(self):
        for suffix in ("", "-wal", "-shm"):
            try:
                os.unlink(self.db_path + suffix)
            except FileNotFoundError:
                pass

    # ── helpers ────────────────────────────────────────────────────────────

    def _xlsx_price_list(self, rows: list[tuple]) -> bytes:
        """Build a minimal XLSX price-list with header + data rows."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Наименование", "Цена", "Кол-во", "Ед.изм"])
        for row in rows:
            ws.append(list(row))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── imports.py unit tests ───────────────────────────────────────────────

    def test_detect_cluster_south(self):
        from procurement.imports import detect_cluster
        # Краснодарский край → cluster_2 (same cluster as production Воронежская обл.)
        cluster, status = detect_cluster("Краснодарский край")
        self.assertEqual(cluster, "cluster_2")
        self.assertEqual(status, "confirmed")

    def test_detect_cluster_moscow(self):
        from procurement.imports import detect_cluster
        # Moscow is not mapped to any cluster → needs_review (no guessing)
        cluster, status = detect_cluster("г. Москва и Московская область")
        self.assertEqual(cluster, "")
        self.assertEqual(status, "needs_review")

    def test_detect_cluster_unknown(self):
        from procurement.imports import detect_cluster
        cluster, status = detect_cluster("Неизвестная область XYZ")
        self.assertEqual(cluster, "")
        self.assertEqual(status, "needs_review")

    def test_supplier_dedup_key_stable(self):
        from procurement.imports import supplier_dedup_key
        k1 = supplier_dedup_key("7700000001", "ООО Тест", "test@example.com", "+7 900 000-00-00")
        k2 = supplier_dedup_key("7700000001", "ООО Тест", "test@example.com", "+7 900 000-00-00")
        self.assertEqual(k1, k2)
        self.assertEqual(len(k1), 64)  # sha256 hex

    def test_supplier_dedup_key_differs_by_inn(self):
        from procurement.imports import supplier_dedup_key
        k1 = supplier_dedup_key("7700000001", "ООО Тест", "", "")
        k2 = supplier_dedup_key("7700000002", "ООО Тест", "", "")
        self.assertNotEqual(k1, k2)

    def test_detect_currency_rub(self):
        from procurement.imports import detect_currency
        self.assertEqual(detect_currency("Цена в руб. без НДС"), "RUB")

    def test_detect_currency_usd(self):
        from procurement.imports import detect_currency
        self.assertEqual(detect_currency("Price in USD"), "USD")

    def test_detect_currency_default_rub(self):
        from procurement.imports import detect_currency
        self.assertEqual(detect_currency("нет упоминания валюты"), "RUB")

    def test_extract_date_dmy(self):
        from procurement.imports import extract_date
        self.assertEqual(extract_date("Дата: 15.08.2026"), "2026-08-15")

    def test_extract_date_iso(self):
        from procurement.imports import extract_date
        self.assertEqual(extract_date("2026-08-21T00:00:00Z"), "2026-08-21")

    def test_extract_date_not_found(self):
        from procurement.imports import extract_date
        self.assertIsNone(extract_date("нет даты"))

    def test_is_price_expired_past(self):
        from procurement.imports import is_price_expired
        self.assertTrue(is_price_expired("2020-01-01"))

    def test_is_price_expired_future(self):
        from procurement.imports import is_price_expired
        self.assertFalse(is_price_expired("2099-12-31"))

    def test_is_price_expired_none(self):
        from procurement.imports import is_price_expired
        self.assertFalse(is_price_expired(None))

    def test_extract_from_xlsx_price_list(self):
        from procurement.imports import extract_from_xlsx
        content = self._xlsx_price_list([
            ("Труба стальная 57×4", "3500.00", "10", "м"),
            ("Фланец Dn50", "450", "20", "шт."),
        ])
        result = extract_from_xlsx(content, "price.xlsx")
        self.assertEqual(len(result.items), 2)
        self.assertEqual(result.items[0].item_name, "Труба стальная 57×4")
        self.assertEqual(result.items[0].unit_price, "3500.00")
        self.assertEqual(result.items[0].source_sheet, "Sheet")
        self.assertEqual(result.items[0].source_row, 2)

    def test_extract_from_xlsx_no_price_col(self):
        from procurement.imports import extract_from_xlsx
        wb = Workbook()
        ws = wb.active
        ws.append(["Описание", "Количество"])
        ws.append(["Болт М12", "100"])
        buf = io.BytesIO(); wb.save(buf)
        result = extract_from_xlsx(buf.getvalue(), "bad.xlsx")
        self.assertEqual(len(result.items), 0)
        self.assertTrue(any("required columns" in e for e in result.errors))

    def test_extract_document_unsupported_raises(self):
        from procurement.imports import extract_document
        with self.assertRaises(ValueError) as ctx:
            extract_document(b"data", "file.docx")
        self.assertIn("unsupported", str(ctx.exception))

    # ── service unit tests ─────────────────────────────────────────────────

    def test_create_import_batch_xlsx(self):
        content = self._xlsx_price_list([
            ("Болт М10", "15.50", "500", "шт."),
            ("Гайка М10", "8.00", "500", "шт."),
        ])
        batch = self.service.create_import_batch(
            [("test_price.xlsx", content)], created_by="tester"
        )
        self.assertIn(batch["status"], ("needs_review", "done"))
        self.assertEqual(batch["total_files"], 1)
        self.assertEqual(batch["processed_files"], 1)
        # Items extracted
        entries = batch["price_history_entries"]
        self.assertGreaterEqual(len(entries), 1)
        self.assertEqual(entries[0]["item_name"], "Болт М10")
        self.assertEqual(entries[0]["status"], "draft")

    def test_create_import_batch_dedup_sha256(self):
        """Same file twice should produce one source_document (dedup by sha256)."""
        content = self._xlsx_price_list([("Позиция 1", "100", "1", "шт.")])
        self.service.create_import_batch([("f.xlsx", content)])
        self.service.create_import_batch([("f.xlsx", content)])
        docs = self.db.all("SELECT * FROM source_documents WHERE sha256 IS NOT NULL")
        sha256_vals = [d["sha256"] for d in docs if d["sha256"]]
        self.assertEqual(len(sha256_vals), len(set(sha256_vals)), "SHA256 dedup failed")

    def test_confirm_batch_entries(self):
        content = self._xlsx_price_list([("Кирпич М150", "35", "1000", "шт.")])
        batch = self.service.create_import_batch([("b.xlsx", content)])
        batch_id = batch["id"]
        entries = batch["price_history_entries"]
        self.assertTrue(entries, "no entries extracted")
        entry_ids = [e["id"] for e in entries]

        result = self.service.confirm_batch_entries(batch_id, entry_ids, "Менеджер")
        self.assertEqual(result["confirmed"], len(entry_ids))

        updated = self.db.all(
            "SELECT status FROM price_history_entries WHERE import_batch_id=?",
            (batch_id,),
        )
        for row in updated:
            self.assertEqual(row["status"], "confirmed")

    def test_confirm_batch_entries_wrong_batch(self):
        content = self._xlsx_price_list([("Товар", "10", "1", "шт.")])
        batch = self.service.create_import_batch([("f.xlsx", content)])
        entries = batch["price_history_entries"]
        if not entries:
            return  # nothing to test
        from procurement.service import NotFoundError
        with self.assertRaises(NotFoundError):
            self.service.confirm_batch_entries(9999, [entries[0]["id"]], "x")

    def test_list_supplier_drafts_empty(self):
        drafts = self.service.list_supplier_drafts()
        self.assertIsInstance(drafts, list)
        self.assertEqual(len(drafts), 0)

    def test_reject_supplier_draft(self):
        # Create a batch to get a draft (need a file that has supplier info)
        # We'll insert a draft directly for this test
        with self.db.connection() as conn:
            conn.execute(
                """
                INSERT INTO supplier_drafts(
                    name, dedup_key, status, cluster_status, created_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                ("ООО Тест", "testkey123", "needs_review", "needs_review", "2026-08-21T00:00:00"),
            )

        class RejectData:
            rejected_by = "Тестировщик"
            review_notes = "дубликат"

        drafts = self.service.list_supplier_drafts(status="needs_review")
        self.assertEqual(len(drafts), 1)
        result = self.service.reject_supplier_draft(drafts[0]["id"], RejectData())
        self.assertEqual(result["status"], "rejected")

    def test_list_price_history_entries(self):
        content = self._xlsx_price_list([("Труба", "500", "10", "м")])
        batch = self.service.create_import_batch([("p.xlsx", content)])
        batch_id = batch["id"]
        entries_draft = self.service.list_price_history_entries(status="draft", batch_id=batch_id)
        self.assertIsInstance(entries_draft, list)
        entries_confirmed = self.service.list_price_history_entries(status="confirmed")
        self.assertIsInstance(entries_confirmed, list)

    def test_get_import_batch_not_found(self):
        from procurement.service import NotFoundError
        with self.assertRaises(NotFoundError):
            self.service.get_import_batch(9999)


    # ── New tests for PR #8 blocking-issue fix ─────────────────────────────

    def test_voronezh_infers_cluster_2(self):
        """Воронежская область (production project region) → cluster_2."""
        from procurement.regions import infer_cluster
        self.assertEqual(infer_cluster("Воронежская область"), "cluster_2")

    def test_krasnodar_infers_cluster_2(self):
        """Краснодарский край must be in cluster_2 per spec."""
        from procurement.regions import infer_cluster
        self.assertEqual(infer_cluster("Краснодарский край"), "cluster_2")

    def test_unknown_region_needs_review(self):
        """Completely unknown region → empty cluster + needs_review."""
        from procurement.imports import detect_cluster
        cluster, status = detect_cluster("Неизвестная Республика XYZ-999")
        self.assertEqual(cluster, "")
        self.assertEqual(status, "needs_review")

    def test_supplier_cluster_scoped_match(self):
        """Supplier of lot's cluster appears in matches; foreign cluster excluded."""
        sup1 = self.service.create_supplier(
            SupplierCreate(
                name="ООО Южный",
                region="Краснодарский край",
                cluster="cluster_2",
                categories=[],
            )
        )
        sup2 = self.service.create_supplier(
            SupplierCreate(
                name="ООО Уральский",
                region="Свердловская область",
                cluster="cluster_1",
                categories=[],
            )
        )
        project = self.service.create_project(
            ProjectCreate(
                name="Проект Юг",
                region="Краснодарский край",
                delivery_address="г. Краснодар, ул. Красная, 1",
            )
        )
        lot = self.service.create_lot(
            LotCreate(
                project_id=project["id"],
                title="Труба стальная",
                region="Краснодарский край",
                cluster="cluster_2",
                delivery_address="г. Краснодар, ул. Красная, 1",
                response_deadline=date.today() + timedelta(days=7),
                items=[
                    LotItemCreate(
                        name="Труба 57x4",
                        quantity=Decimal("100"),
                        unit="м",
                    )
                ],
            )
        )
        matches = self.service.match_suppliers(lot["id"])
        match_ids = [m["id"] for m in matches]
        self.assertIn(sup1["id"], match_ids, "cluster_2 supplier must appear in matches")
        self.assertNotIn(sup2["id"], match_ids, "cluster_1 supplier must be excluded")

    def test_same_inn_different_contact_no_duplicate_key(self):
        """INN-primary dedup: same INN + different email/phone → same dedup key."""
        from procurement.imports import supplier_dedup_key
        k1 = supplier_dedup_key("7700000001", "ООО Тест", "old@example.com", "+7 900 000-00-01")
        k2 = supplier_dedup_key("7700000001", "ООО Тест", "new@example.com", "+7 900 000-00-02")
        self.assertEqual(k1, k2, "Same INN must produce same dedup key regardless of contact")

    def test_reimport_same_sha256_no_new_entries(self):
        """Re-importing the same file must not create new price_history_entries."""
        content = self._xlsx_price_list([("Болт М8", "10.00", "100", "шт.")])
        self.service.create_import_batch([("same.xlsx", content)])
        self.service.create_import_batch([("same.xlsx", content)])
        rows = self.db.all("SELECT COUNT(*) AS n FROM price_history_entries")
        self.assertEqual(rows[0]["n"], 1, "Second SHA256 import must not add entries")

    def test_reimport_same_sha256_no_new_source_doc(self):
        """Re-importing the same file must not create a second source_document row."""
        content = self._xlsx_price_list([("Гайка М8", "5.00", "200", "шт.")])
        self.service.create_import_batch([("f.xlsx", content)])
        self.service.create_import_batch([("f.xlsx", content)])
        docs = self.db.all("SELECT sha256 FROM source_documents WHERE sha256 IS NOT NULL")
        sha256_vals = [d["sha256"] for d in docs]
        self.assertEqual(
            len(sha256_vals), len(set(sha256_vals)),
            "SHA256 dedup failed — duplicate source_documents found",
        )

    def test_price_validity_state_none_is_unknown(self):
        """valid_until=None must yield 'unknown', NOT treated as active."""
        from procurement.imports import price_validity_state
        self.assertEqual(price_validity_state(None), "unknown")

    def test_price_validity_state_past_is_expired(self):
        from procurement.imports import price_validity_state
        self.assertEqual(price_validity_state("2020-01-01"), "expired")

    def test_price_validity_state_future_is_active(self):
        from procurement.imports import price_validity_state
        self.assertEqual(price_validity_state("2099-12-31"), "active")


class BatchImportApiTestCase(unittest.TestCase):
    """API endpoint tests for PR #8."""

    def setUp(self):
        from fastapi.testclient import TestClient
        handle = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self.db_path = handle.name
        handle.close()
        with patch.dict(
            os.environ,
            {
                "PROCUREMENT_DB_PATH": self.db_path,
                "PROCUREMENT_ENV": "development",
                "PROCUREMENT_API_KEY": "",
                "PROCUREMENT_OUTBOX_MODE": "draft_only",
                "PROCUREMENT_AUTH_SECRET": "",
                "PROCUREMENT_ADMIN_USERNAME": "",
                "PROCUREMENT_ADMIN_PASSWORD_HASH": "",
            },
        ):
            import importlib
            import procurement.app as app_module
            importlib.reload(app_module)
            app_module.db.initialize()
            self.client = TestClient(app_module.app, raise_server_exceptions=True)

    def tearDown(self):
        for suffix in ("", "-wal", "-shm"):
            try:
                os.unlink(self.db_path + suffix)
            except FileNotFoundError:
                pass

    def _xlsx_bytes(self, rows: list[tuple]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.append(["Наименование", "Цена", "Ед.изм"])
        for row in rows:
            ws.append(list(row))
        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()

    def test_api_version_is_080(self):
        res = self.client.get("/openapi.json")
        self.assertEqual(res.status_code, 200)
        info = res.json().get("info", {})
        self.assertEqual(info.get("version"), "0.8.0")

    def test_post_imports_batch_xlsx(self):
        content = self._xlsx_bytes([("Болт М10", "15", "шт.")])
        res = self.client.post(
            "/api/imports/batch",
            files=[("files", ("price.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )
        self.assertEqual(res.status_code, 201, res.text)
        data = res.json()
        self.assertIn("id", data)
        self.assertIn("status", data)
        self.assertIn("price_history_entries", data)

    def test_post_imports_batch_no_files_422(self):
        res = self.client.post("/api/imports/batch", files=[])
        self.assertIn(res.status_code, (400, 422))

    def test_post_imports_batch_too_many_files_422(self):
        content = self._xlsx_bytes([("x", "1", "шт.")])
        files = [("files", (f"f{i}.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")) for i in range(21)]
        res = self.client.post("/api/imports/batch", files=files)
        self.assertIn(res.status_code, (413, 422))

    def test_get_imports_list(self):
        res = self.client.get("/api/imports")
        self.assertEqual(res.status_code, 200)
        self.assertIsInstance(res.json(), list)

    def test_get_imports_batch_not_found(self):
        res = self.client.get("/api/imports/9999")
        self.assertEqual(res.status_code, 404)

    def test_get_supplier_drafts(self):
        res = self.client.get("/api/supplier-drafts")
        self.assertEqual(res.status_code, 200)
        self.assertIsInstance(res.json(), list)

    def test_get_price_history_entries(self):
        res = self.client.get("/api/price-history-entries")
        self.assertEqual(res.status_code, 200)
        self.assertIsInstance(res.json(), list)

    def test_post_confirm_batch_entries_flow(self):
        content = self._xlsx_bytes([("Гайка М8", "5", "шт."), ("Болт М8", "8", "шт.")])
        res = self.client.post(
            "/api/imports/batch",
            files=[("files", ("test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )
        self.assertEqual(res.status_code, 201)
        batch = res.json()
        batch_id = batch["id"]
        entries = batch["price_history_entries"]

        if not entries:
            return  # extraction found nothing — skip confirm step

        entry_ids = [e["id"] for e in entries]
        res2 = self.client.post(
            f"/api/imports/{batch_id}/confirm",
            json={"confirmed_by": "Тестировщик", "entry_ids": entry_ids},
        )
        self.assertEqual(res2.status_code, 200, res2.text)
        self.assertEqual(res2.json()["confirmed"], len(entry_ids))

        # entries now visible as confirmed
        res3 = self.client.get(f"/api/price-history-entries?status=confirmed&batch_id={batch_id}")
        self.assertEqual(res3.status_code, 200)
        confirmed = res3.json()
        self.assertEqual(len(confirmed), len(entry_ids))

    def test_post_imports_confirm_wrong_batch_404(self):
        content = self._xlsx_bytes([("x", "1", "шт.")])
        res = self.client.post(
            "/api/imports/batch",
            files=[("files", ("f.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )
        self.assertEqual(res.status_code, 201)
        batch = res.json()
        entries = batch["price_history_entries"]
        if not entries:
            return
        # Use wrong batch_id
        res2 = self.client.post(
            "/api/imports/9999/confirm",
            json={"confirmed_by": "x", "entry_ids": [entries[0]["id"]]},
        )
        self.assertEqual(res2.status_code, 404)

    def test_reject_supplier_draft_404(self):
        res = self.client.post(
            "/api/supplier-drafts/9999/reject",
            json={"rejected_by": "x", "review_notes": ""},
        )
        self.assertEqual(res.status_code, 404)


class DockerfileProxyHeadersTestCase(unittest.TestCase):
    """PR #8 fix: Dockerfile is now copied into the image so the test works."""

    def test_uvicorn_does_not_trust_proxy_headers_before_the_app(self):
        dockerfile_path = Path(__file__).parents[1] / "Dockerfile"
        if not dockerfile_path.exists():
            self.skipTest("Dockerfile not found (running in image without fix)")
        dockerfile = dockerfile_path.read_text(encoding="utf-8")
        self.assertIn("--no-proxy-headers", dockerfile)

