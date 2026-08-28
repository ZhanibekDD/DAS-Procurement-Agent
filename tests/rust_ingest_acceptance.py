from __future__ import annotations

import json
import os
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


class RustIngestAcceptanceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        binary = os.environ.get("DAS_INGEST_TEST_BINARY", "").strip()
        if not binary:
            raise RuntimeError("DAS_INGEST_TEST_BINARY is required")
        cls.binary = Path(binary).resolve()
        if not cls.binary.is_file():
            raise RuntimeError("DAS_INGEST_TEST_BINARY does not exist")

    def run_ingest(self, path: Path) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [
                str(self.binary),
                "--input",
                str(path),
                "--filename",
                path.name,
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
        )

    def test_xlsx_extracts_supplier_totals_and_traceable_items(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "comparison.xlsx"
            workbook = Workbook()
            summary = workbook.active
            summary.title = "Сводная"
            summary.append(["КП Электрика"])
            summary.append(["Подрядчик", "ООО Первый", "ТОО Второй"])
            summary.append(["Цена работ", 100000, 90000])
            summary.append(["Цена материалов по проекту", 250000, 240000])
            summary.append(["НДС", "НДС 12%", "без НДС"])
            summary.append(["Итого по проекту", 350000, 330000])
            items = workbook.create_sheet("Кровля")
            items.append(["ООО Первый"])
            items.append(
                [
                    "№",
                    "Наименование",
                    "Ед. изм.",
                    "Количество",
                    "Цена за ед.",
                    "Сумма",
                ]
            )
            items.append(
                [1, "Обеспыливание поверхности", "м2", 1134.26, 66, 74861.16]
            )
            workbook.save(path)

            completed = self.run_ingest(path)
            self.assertEqual(completed.returncode, 0, completed.stderr)
            payload = json.loads(completed.stdout)
            self.assertEqual(payload["schema_version"], "1.0")
            self.assertEqual(len(payload["sheets"]), 2)
            self.assertEqual(len(payload["suppliers"]), 2)
            self.assertEqual(len(payload["summaries"]), 2)
            self.assertEqual(payload["summaries"][0]["total_project"], 350000)
            self.assertEqual(len(payload["line_items"]), 1)
            self.assertEqual(payload["line_items"][0]["source_row"], 3)
            self.assertEqual(payload["line_items"][0]["quantity"], 1134.26)
            self.assertEqual(len(payload["source"]["sha256"]), 64)

    def test_csv_keeps_first_data_row(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "offer.csv"
            path.write_text(
                "Наименование;Ед. изм.;Количество;Цена за ед.;Сумма\n"
                "Кабель;м;100;25,5;2550\n",
                encoding="utf-8",
            )
            completed = self.run_ingest(path)
            self.assertEqual(completed.returncode, 0, completed.stderr)
            payload = json.loads(completed.stdout)
            self.assertEqual(len(payload["line_items"]), 1)
            self.assertEqual(payload["line_items"][0]["name"], "Кабель")

    def test_xlsx_maps_multi_supplier_price_columns(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "supplier_matrix.xlsx"
            workbook = Workbook()
            summary = workbook.active
            summary.title = "Сводная"
            summary.append(["КП Электрика"])
            summary.append(["Подрядчик", "ООО Первый", "ТОО Второй"])
            summary.append(["Итого", 3700, 2300])

            matrix = workbook.create_sheet("Электрика")
            matrix.append([None, None, None, None, "ООО Первый", None, None, None, "ТОО Второй"])
            matrix.append(
                [
                    "Наименование",
                    "Ед. изм.",
                    "Количество",
                    None,
                    "Работы",
                    None,
                    "Материалы",
                    None,
                    "Работы",
                    None,
                ]
            )
            matrix.append(
                [
                    None,
                    None,
                    None,
                    None,
                    "Цена за ед.",
                    "Сумма",
                    "Цена за ед.",
                    "Сумма",
                    "Цена за ед.",
                    "Сумма",
                ]
            )
            matrix.append(["Кабель", "м", 100, None, 25, 2500, 12, 1200, 23, 2300])
            matrix.append(["Всего с НДС", None, None, None, None, 2500, None, 1200, None, 2300])
            workbook.save(path)

            completed = self.run_ingest(path)
            self.assertEqual(completed.returncode, 0, completed.stderr)
            payload = json.loads(completed.stdout)
            self.assertEqual(len(payload["line_items"]), 1)
            item = payload["line_items"][0]
            self.assertEqual(len(item["offers"]), 3)
            self.assertEqual(item["offers"][0]["supplier"], "ООО Первый")
            self.assertEqual(item["offers"][0]["category"], "work")
            self.assertEqual(item["offers"][0]["unit_price"], 25)
            self.assertEqual(item["offers"][0]["source_unit_price_column"], 5)
            self.assertEqual(item["offers"][1]["category"], "materials_project")
            self.assertEqual(item["offers"][2]["supplier"], "ТОО Второй")
            self.assertEqual(payload["warnings"], [])

    def test_invalid_xlsx_fails_closed(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "broken.xlsx"
            path.write_bytes(b"not-a-zip")
            completed = self.run_ingest(path)
            self.assertNotEqual(completed.returncode, 0)
            self.assertIn("invalid XLSX payload", completed.stderr)


if __name__ == "__main__":
    unittest.main()
