from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import SupplierCreate


HEADER_ALIASES = {
    "name": {"наименование", "поставщик", "контрагент", "компания", "name", "supplier"},
    "tax_id": {"инн", "бин", "иин", "огрн", "tax_id", "tax id"},
    "region": {"регион", "город", "область", "region", "city"},
    "email": {"email", "e-mail", "электронная почта", "почта"},
    "phone": {"телефон", "phone", "мобильный"},
    "telegram": {"telegram", "телеграм", "tg"},
    "max_contact": {"max", "мах", "макс", "max контакт", "мах контакт"},
    "cluster": {"кластер", "cluster"},
    "categories": {"категория", "категории", "товары", "услуги", "category"},
    "rating": {"рейтинг", "rating", "оценка"},
    "verified": {"проверен", "проверенный", "verified"},
}


@dataclass
class ImportPreview:
    rows: list[SupplierCreate] = field(default_factory=list)
    errors: list[dict[str, object]] = field(default_factory=list)
    headers: list[str] = field(default_factory=list)


def _normalize(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("ё", "е").split())


def _header_mapping(headers: Iterable[object]) -> tuple[list[str], dict[int, str]]:
    originals = [str(value or "").strip() for value in headers]
    mapping: dict[int, str] = {}
    for index, header in enumerate(originals):
        normalized = _normalize(header)
        for field_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapping[index] = field_name
                break
    if "name" not in mapping.values():
        raise ValueError("supplier table must contain a supplier/name column")
    if "region" not in mapping.values():
        raise ValueError("supplier table must contain a region/city column")
    return originals, mapping


def _bool(value: object) -> bool:
    return _normalize(value) in {"1", "true", "yes", "да", "проверен", "проверенный"}


def _supplier_from_row(row: list[object], mapping: dict[int, str]) -> SupplierCreate | None:
    values = {field_name: row[index] if index < len(row) else "" for index, field_name in mapping.items()}
    if not any(str(value or "").strip() for value in values.values()):
        return None
    raw_categories = str(values.get("categories") or "")
    categories = [part.strip() for part in raw_categories.replace(";", ",").split(",") if part.strip()]
    raw_rating = str(values.get("rating") or "").replace(",", ".").strip()
    return SupplierCreate(
        name=str(values.get("name") or ""),
        tax_id=str(values.get("tax_id") or ""),
        region=str(values.get("region") or ""),
        email=str(values.get("email") or ""),
        phone=str(values.get("phone") or ""),
        telegram=str(values.get("telegram") or ""),
        max_contact=str(values.get("max_contact") or ""),
        cluster=str(values.get("cluster") or ""),
        categories=categories,
        rating=float(raw_rating) if raw_rating else 3.0,
        verified=_bool(values.get("verified")),
    )


def _parse_rows(rows: Iterable[tuple[object, ...]]) -> ImportPreview:
    iterator = iter(rows)
    try:
        header_row = next(iterator)
    except StopIteration as exc:
        raise ValueError("supplier table is empty") from exc
    headers, mapping = _header_mapping(header_row)
    preview = ImportPreview(headers=headers)
    for row_number, row in enumerate(iterator, start=2):
        try:
            supplier = _supplier_from_row(list(row), mapping)
            if supplier is not None:
                preview.rows.append(supplier)
        except Exception as exc:
            preview.errors.append({"row": row_number, "error": str(exc)})
    return preview


def parse_supplier_table(content: bytes, filename: str) -> ImportPreview:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        return _parse_rows(tuple(row) for row in csv.reader(io.StringIO(text), dialect))
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            return _parse_rows(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
    raise ValueError("only .csv and .xlsx supplier tables are supported")


# ── PR #8: price-list / КП / счёт batch extraction ───────────────────────────

import hashlib
import re
from dataclasses import dataclass, field
from datetime import date as _date


# ---------------------------------------------------------------------------
# Cluster detection by region keyword
# ---------------------------------------------------------------------------
_REGION_CLUSTER: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r'краснодар|ростов|ставропол|адыге|калмык', re.I), 'cluster_south'),
    (re.compile(r'москв|подмосков|московск', re.I), 'cluster_moscow'),
    (re.compile(r'санкт.петербург|ленинград|питер', re.I), 'cluster_spb'),
    (re.compile(r'екатеринбург|свердлов|урал', re.I), 'cluster_ural'),
    (re.compile(r'новосибирск|томск|омск|кемеров|алтай', re.I), 'cluster_siberia'),
    (re.compile(r'красноярск|иркутск|якут|бурят', re.I), 'cluster_east_siberia'),
    (re.compile(r'владивосток|хабаровск|приморск|сахалин', re.I), 'cluster_far_east'),
    (re.compile(r'казань|татарстан|башкортостан|самар|уфа|пермь', re.I), 'cluster_volga'),
    (re.compile(r'воронеж|белгород|курск|орёл|орел|липецк|тамбов', re.I), 'cluster_central_black'),
]


def detect_cluster(region: str) -> tuple[str, str]:
    """Return (cluster, cluster_status) for a region string."""
    for pattern, cluster in _REGION_CLUSTER:
        if pattern.search(region):
            return cluster, 'confirmed'
    return '', 'needs_review'


# ---------------------------------------------------------------------------
# Deduplication key
# ---------------------------------------------------------------------------
def supplier_dedup_key(tax_id: str, name: str, email: str, phone: str) -> str:
    raw = '|'.join([
        tax_id.strip(),
        re.sub(r'[^а-яa-z0-9]+', ' ', name.casefold()).strip(),
        email.casefold().strip(),
        re.sub(r'[^0-9]', '', phone),
    ])
    return hashlib.sha256(raw.encode()).hexdigest()


# ---------------------------------------------------------------------------
# Currency normalisation
# ---------------------------------------------------------------------------
_CURRENCY_MAP = {
    'руб': 'RUB', 'rub': 'RUB', 'rur': 'RUB', 'р.': 'RUB',
    'usd': 'USD', 'доллар': 'USD', '$': 'USD',
    'eur': 'EUR', 'евро': 'EUR', '€': 'EUR',
}


def detect_currency(text: str) -> str:
    t = text.casefold()
    for k, v in _CURRENCY_MAP.items():
        if k in t:
            return v
    return 'RUB'


# ---------------------------------------------------------------------------
# Date extraction
# ---------------------------------------------------------------------------
_DATE_RE = re.compile(
    r'(\d{1,2})[./\-](\d{1,2})[./\-](\d{2,4})'
    r'|(\d{4})-(\d{2})-(\d{2})'
)


def extract_date(text: str) -> str | None:
    m = _DATE_RE.search(text)
    if not m:
        return None
    if m.group(4):            # YYYY-MM-DD
        y, mo, d = m.group(4), m.group(5), m.group(6)
    else:                     # D.M.Y or D/M/Y
        d, mo, y = m.group(1), m.group(2), m.group(3)
    if len(y) == 2:
        y = '20' + y
    try:
        return _date(int(y), int(mo), int(d)).isoformat()
    except ValueError:
        return None


def is_price_expired(valid_until: str | None) -> bool:
    if not valid_until:
        return False
    try:
        return _date.fromisoformat(valid_until) < _date.today()
    except ValueError:
        return False


# ---------------------------------------------------------------------------
# Column detection helpers
# ---------------------------------------------------------------------------
_PRICE_COL_NAMES: set[str] = {
    'price', 'цена', 'стоимость', 'цена с ндс', 'цена без ндс',
    'прайс', 'ед', 'за ед', 'unit price', 'unit_price', 'цена ед', 'цена/ед',
}
_NAME_COL_NAMES: set[str] = {
    'наименование', 'название', 'товар', 'позиция', 'item', 'name', 'description',
    'наим.', 'описание', 'продукция', 'материал', 'номенклатура',
}
_QTY_COL_NAMES: set[str] = {
    'кол-во', 'количество', 'qty', 'quantity', 'кол.', 'объём', 'объем', 'кол',
}
_UNIT_COL_NAMES: set[str] = {
    'ед.изм', 'ед. изм', 'единица', 'unit', 'ед', 'uom', 'ед.изм.',
}


def _col_index(headers: list[str], names: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if h.casefold().strip() in names:
            return i
    for i, h in enumerate(headers):
        hcf = h.casefold().strip()
        if any(n in hcf for n in names):
            return i
    return None


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------
@dataclass
class ExtractedItem:
    item_name: str
    normalized_name: str
    brand: str
    quantity: str
    unit: str
    unit_price: str
    total_price: str
    currency: str
    vat_included: bool
    source_page: int | None
    source_sheet: str
    source_row: int | None
    source_cell: str
    source_text: str


@dataclass
class DocumentExtractResult:
    filename: str
    sha256: str
    document_type: str        # price_list | invoice | commercial_offer | unknown
    supplier_name: str
    supplier_tax_id: str
    supplier_region: str
    supplier_email: str
    supplier_phone: str
    supplier_contact: str
    document_date: str | None
    valid_until: str | None
    currency: str
    vat_included: bool
    items: list[ExtractedItem] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Document type classifier
# ---------------------------------------------------------------------------
def _classify_document(text: str) -> str:
    tl = text.casefold()
    if any(w in tl for w in ('счёт-фактура', 'счет-фактура', 'упд')):
        return 'invoice'
    if any(w in tl for w in ('коммерческое предложение', 'кп №', 'кп от')):
        return 'commercial_offer'
    if any(w in tl for w in ('прайс-лист', 'price list', 'прайс лист', 'прайслист')):
        return 'price_list'
    if any(w in tl for w in ('счёт №', 'счет №', 'счёт на', 'счет на')):
        return 'invoice'
    if any(w in tl for w in ('предложение', 'спецификация')):
        return 'commercial_offer'
    return 'unknown'


# supplier regex patterns
_SUPPLIER_INN_RE = re.compile(r'ИНН[:\s]+([0-9]{10,12})', re.I)
_SUPPLIER_EMAIL_RE = re.compile(r'[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}')
_SUPPLIER_PHONE_RE = re.compile(r'(?:\+7|8)[\s\-]?\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}')
_ORG_RE = re.compile(r'(?:ООО|ИП|АО|ЗАО|ПАО)[\s"«]+([^»"\n]{3,80})', re.I)


# ---------------------------------------------------------------------------
# PDF extraction
# ---------------------------------------------------------------------------
def _extract_pdf_text(content: bytes) -> tuple[list[str], list[str]]:
    """Return (page_texts, errors)."""
    try:
        from pypdf import PdfReader
        import io as _io
        reader = PdfReader(_io.BytesIO(content))
        pages = [page.extract_text() or '' for page in reader.pages]
        return pages, []
    except Exception as exc:
        return [], [f'PDF parse error: {exc}']


def _extract_items_from_pdf_text(
    pages: list[str], currency: str, vat_included: bool
) -> list[ExtractedItem]:
    items: list[ExtractedItem] = []
    # Heuristic: lines with item name + qty + price at end
    price_re = re.compile(r'^(.{5,80?})\s+(\d[\d\s.,]{0,14})\s+(\d[\d\s.,]*[.,]\d{2})\s*$')
    for page_num, text in enumerate(pages, start=1):
        for row_num, line in enumerate(text.splitlines(), start=1):
            line = line.strip()
            if len(line) < 10:
                continue
            m = price_re.match(line)
            if not m:
                continue
            name_raw = m.group(1).strip()
            qty_raw = m.group(2).strip()
            price_raw = m.group(3).replace('\xa0', '').replace(' ', '')
            norm = re.sub(r'[^а-яa-z0-9 ]+', ' ', name_raw.casefold()).strip()
            items.append(ExtractedItem(
                item_name=name_raw,
                normalized_name=norm,
                brand='',
                quantity=qty_raw,
                unit='шт.',
                unit_price=price_raw,
                total_price='',
                currency=currency,
                vat_included=vat_included,
                source_page=page_num,
                source_sheet='',
                source_row=row_num,
                source_cell='',
                source_text=line,
            ))
    return items


def extract_from_pdf(content: bytes, filename: str) -> DocumentExtractResult:
    sha256 = hashlib.sha256(content).hexdigest()
    pages, errors = _extract_pdf_text(content)
    full_text = '\n'.join(pages)
    header_text = '\n'.join(pages[:2]) if pages else ''

    doc_type = _classify_document(full_text)
    currency = detect_currency(full_text)
    vat_included = 'без ндс' not in full_text.casefold()
    doc_date = extract_date(full_text)

    valid_re = re.compile(
        r'(?:действует\s+до'
        r'|действительно\s+до'
        r'|срок\s+действия)'
        r'[^\d]*(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})', re.I
    )
    valid_m = valid_re.search(full_text)
    valid_until = extract_date(valid_m.group(1)) if valid_m else None

    tax_id_m = _SUPPLIER_INN_RE.search(header_text)
    email_m = _SUPPLIER_EMAIL_RE.search(header_text)
    phone_m = _SUPPLIER_PHONE_RE.search(header_text)
    org_m = _ORG_RE.search(header_text)
    supplier_name = org_m.group(0).strip()[:120] if org_m else ''

    items = _extract_items_from_pdf_text(pages, currency, vat_included)

    return DocumentExtractResult(
        filename=filename,
        sha256=sha256,
        document_type=doc_type,
        supplier_name=supplier_name,
        supplier_tax_id=tax_id_m.group(1) if tax_id_m else '',
        supplier_region='',
        supplier_email=email_m.group(0) if email_m else '',
        supplier_phone=phone_m.group(0) if phone_m else '',
        supplier_contact='',
        document_date=doc_date,
        valid_until=valid_until,
        currency=currency,
        vat_included=vat_included,
        items=items,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# XLSX extraction
# ---------------------------------------------------------------------------
def extract_from_xlsx(content: bytes, filename: str) -> DocumentExtractResult:
    from openpyxl import load_workbook
    import io as _io

    sha256 = hashlib.sha256(content).hexdigest()
    errors: list[str] = []

    try:
        wb = load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return DocumentExtractResult(
            filename=filename, sha256=sha256, document_type='unknown',
            supplier_name='', supplier_tax_id='', supplier_region='',
            supplier_email='', supplier_phone='', supplier_contact='',
            document_date=None, valid_until=None, currency='RUB', vat_included=True,
            items=[], errors=[f'XLSX open error: {exc}'],
        )

    all_items: list[ExtractedItem] = []
    header_text_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        for row in rows[:10]:
            header_text_parts.append(' '.join(str(c) for c in row if c is not None))

        header_idx: int | None = None
        headers: list[str] = []
        for i, row in enumerate(rows):
            row_strs = [str(c).strip() if c is not None else '' for c in row]
            row_cf = ' '.join(row_strs).casefold()
            if any(n in row_cf for n in (
                'наименование',
                'цена', 'price', 'item', 'name',
                'количество',
            )):
                header_idx = i
                headers = row_strs
                break

        if header_idx is None:
            errors.append(f'Sheet {sheet_name!r}: no header row found')
            continue

        name_col = _col_index(headers, _NAME_COL_NAMES)
        price_col = _col_index(headers, _PRICE_COL_NAMES)
        qty_col = _col_index(headers, _QTY_COL_NAMES)
        unit_col = _col_index(headers, _UNIT_COL_NAMES)

        if name_col is None or price_col is None:
            errors.append(f'Sheet {sheet_name!r}: required columns not found; headers={headers[:8]}')
            continue

        max_col = max(c for c in [name_col, price_col, qty_col, unit_col] if c is not None)

        for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            cells = [str(c).strip() if c is not None else '' for c in row]
            if len(cells) <= max_col:
                continue
            name = cells[name_col]
            price = cells[price_col]
            qty = cells[qty_col] if qty_col is not None and qty_col < len(cells) else ''
            unit = cells[unit_col] if unit_col is not None and unit_col < len(cells) else ''
            if not name or not price:
                continue
            price_clean = re.sub(r'[^\d.,]', '', str(price))
            if not price_clean:
                continue
            norm = re.sub(r'[^а-яa-z0-9 ]+', ' ', name.casefold()).strip()
            all_items.append(ExtractedItem(
                item_name=name,
                normalized_name=norm,
                brand='',
                quantity=qty,
                unit=unit,
                unit_price=price_clean,
                total_price='',
                currency='RUB',
                vat_included=True,
                source_page=None,
                source_sheet=sheet_name,
                source_row=row_idx,
                source_cell='',
                source_text='|'.join(cells[:8]),
            ))

    header_text = ' '.join(header_text_parts)
    doc_type = _classify_document(header_text)
    currency = detect_currency(header_text)
    vat_included = 'без ндс' not in header_text.casefold()
    doc_date = extract_date(header_text)

    tax_id_m = _SUPPLIER_INN_RE.search(header_text)
    email_m = _SUPPLIER_EMAIL_RE.search(header_text)
    phone_m = _SUPPLIER_PHONE_RE.search(header_text)
    org_m = _ORG_RE.search(header_text)
    supplier_name = org_m.group(0).strip()[:120] if org_m else ''

    return DocumentExtractResult(
        filename=filename,
        sha256=sha256,
        document_type=doc_type,
        supplier_name=supplier_name,
        supplier_tax_id=tax_id_m.group(1) if tax_id_m else '',
        supplier_region='',
        supplier_email=email_m.group(0) if email_m else '',
        supplier_phone=phone_m.group(0) if phone_m else '',
        supplier_contact='',
        document_date=doc_date,
        valid_until=None,
        currency=currency,
        vat_included=vat_included,
        items=all_items,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# Public dispatcher
# ---------------------------------------------------------------------------
def extract_document(content: bytes, filename: str) -> DocumentExtractResult:
    """Extract items and supplier info from a КП/invoice/price-list file."""
    suffix = Path(filename).suffix.lower()
    if suffix == '.pdf':
        return extract_from_pdf(content, filename)
    if suffix in ('.xlsx', '.xls'):
        return extract_from_xlsx(content, filename)
    raise ValueError(f'unsupported file type for price-list extraction: {suffix!r}')
